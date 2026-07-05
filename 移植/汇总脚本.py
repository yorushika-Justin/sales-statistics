"""
销售汇总脚本 v1.7
================

用途:
    将供应商日销售汇总 xls/xlsx 按"品牌名称"分组汇总销售金额,
    同时按"品名"逐行展示明细,
    输出 2 个 PNG 表格图片.

用法:
    1. 拖到 exe 图标: 把 xls/xlsx 文件拖到打包后的 exe 上, 直接生成 PNG
    2. 双击 exe:     打开拖拽窗口, 把文件拖入窗口中
    3. 命令行:       python 汇总脚本.py <xls/xlsx 路径>

输入:
    sys.argv[1] = xls/xlsx 文件路径 (可选, 不传则启动 GUI 窗口)

输出:
    D:/16678/相册/销售额/<日期>.png        (品牌汇总, 自动打开)
    D:/16678/相册/销售额/<日期>_品名.png   (品名明细, 不自动打开)

日期逻辑:
    xls/xlsx 内部 Row 2 日期范围有滞后性, 结束日期 -1 天为实际最后完整统计日.
    例: "2026-06-05 00:00:00 至 2026-06-06 23:59:59" 实际只统计了 6.5, 显示 "2026-06-05"

处理逻辑:
    1. 定位"品牌名称"列和"销售金额"列 (自动扫描表头)
    2. 过滤非数据行 (空品牌/合计行)
    3. 按品牌 sum 销售金额, float 累加, 保留 2 位小数
    4. 降序排序
    5. Pillow 画表格: 标题 + 表头 + N 行品牌 + 加粗总销售金额
    6. 同时按品名逐行展示明细 (行号/条码/品名/品牌/数量/金额/库存 7 列)
    7. 生成后用系统默认看图工具自动打开品牌汇总 PNG

格式支持:
    - .xls: xlrd 1.2.0 (最后支持 xls 的版本)
    - .xlsx: openpyxl (Excel 2007+ 格式)

异常处理:
    - 缺参数 -> 启动 GUI 拖拽窗口
    - 文件不存在 / 后缀不对 -> tkinter 弹窗
    - 输出目录不存在 -> tkinter 弹窗提示
    - 找不到品牌/金额列 -> 弹窗 + 写日志
    - 任何运行时异常 -> 弹窗 + 写日志到文件同目录的 汇总错误.log
"""
import sys
import os
import re
import traceback
import tkinter as tk
from datetime import datetime, timedelta
from collections import defaultdict

import xlrd
import openpyxl
from PIL import Image, ImageDraw, ImageFont


# ==================== 常量 ====================
OUTPUT_DIR = r"D:\16678\相册\销售额"         # 固定输出目录
FONT_PATH = r"C:\Windows\Fonts\msyh.ttc"   # 微软雅黑
COLOR_BORDER = (60, 60, 60)
COLOR_TEXT = (0, 0, 0)
COLOR_BG = (255, 255, 255)

# 字体大小
SIZE_TITLE = 32
SIZE_HEADER = 26
SIZE_BODY = 22
SIZE_TOTAL = 24

# 表格尺寸 (像素)
COL_BRAND_W = 260
COL_AMOUNT_W = 240
ROW_H = 52
PADDING = 30

# 品名明细表格尺寸 (像素)
DETAIL_COL_NUM_W = 70
DETAIL_COL_BARCODE_W = 200
DETAIL_COL_ITEM_W = 280
DETAIL_COL_BRAND_W = 200
DETAIL_COL_QTY_W = 120
DETAIL_COL_AMOUNT_W = 200
DETAIL_COL_STOCK_W = 120
DETAIL_ROW_H = 52


# ==================== 工具函数 ====================
def extract_date_from_filename(filename: str) -> str:
    """从文件名提取 YYYY-MM-DD 格式日期, 失败则用今天"""
    m = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if m:
        return m.group(1)
    return datetime.now().strftime('%Y-%m-%d')


def show_error(msg: str, log_path: str = None, parent=None):
    """用 tkinter 弹窗显示错误, 必要时写日志"""
    try:
        from tkinter import messagebox
        if parent is not None:
            messagebox.showerror("销售汇总 - 错误", msg, parent=parent)
        else:
            temp_root = tk.Tk()
            temp_root.withdraw()
            messagebox.showerror("销售汇总 - 错误", msg)
            temp_root.destroy()
    except Exception:
        print(f"[错误] {msg}", file=sys.stderr)
    if log_path:
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f'[{datetime.now()}] {msg}\n')
                if sys.exc_info()[0] is not None:
                    f.write(f'{traceback.format_exc()}\n')
        except Exception:
            pass


def launch_gui():
    """双击 exe 时启动拖拽窗口 (需要 tkinterdnd2)"""
    try:
        from tkinterdnd2 import DND_FILES, TkinterDnD
    except ImportError:
        show_error("拖拽窗口组件缺失。\n请将 xls 文件直接拖到 exe 图标上使用。")
        return

    root = TkinterDnD.Tk()
    root.title("销售汇总")
    root.geometry("480x320")
    root.resizable(False, False)
    root.configure(bg="#ffffff")

    # 居中
    root.update_idletasks()
    w, h = root.winfo_reqwidth(), root.winfo_reqheight()
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    root.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

    # 标题
    tk.Label(root, text="销售汇总工具", font=("微软雅黑", 16, "bold"),
             bg="#ffffff", fg="#333333").pack(pady=(30, 10))

    # 拖放区
    drop_frame = tk.Frame(root, bg="#f5f7fa", relief="groove", bd=2,
                          highlightbackground="#c0c0c0", highlightthickness=1)
    drop_frame.pack(fill="both", expand=True, padx=30, pady=(0, 15))

    tk.Label(drop_frame, text="将 .xls / .xlsx 文件拖入此处",
             font=("微软雅黑", 13), bg="#f5f7fa", fg="#888888").pack(expand=True)

    tk.Label(drop_frame, text="也可将文件直接拖到 exe 图标上",
             font=("微软雅黑", 9), bg="#f5f7fa", fg="#aaaaaa").pack(pady=(0, 15))

    # 状态栏
    status_var = tk.StringVar()
    tk.Label(root, textvariable=status_var, font=("微软雅黑", 10),
             bg="#ffffff", fg="#28a745").pack(pady=(0, 15))

    def on_drop(event):
        paths = root.tk.splitlist(event.data)
        if not paths:
            return
        path = paths[0].strip()
        if not (path.lower().endswith('.xls') or path.lower().endswith('.xlsx')):
            status_var.set("仅支持 .xls / .xlsx 文件")
            return
        if len(paths) > 1:
            status_var.set(f"检测到 {len(paths)} 个文件，仅处理第 1 个")
            root.update_idletasks()
        status_var.set("处理中……")
        root.update_idletasks()
        xls_path = os.path.abspath(path)
        log_path = os.path.join(os.path.dirname(xls_path) or '.', '汇总错误.log')
        is_xlsx = xls_path.lower().endswith('.xlsx')
        try:
            if not os.path.exists(xls_path):
                show_error(f"文件不存在:\n{xls_path}", log_path, parent=root)
                status_var.set("")
                return
            if not os.path.isdir(OUTPUT_DIR):
                show_error(f"输出目录不存在:\n{OUTPUT_DIR}\n请先创建该目录。", log_path, parent=root)
                status_var.set("")
                return
            data, date_str = read_xlsx(xls_path, log_path=log_path) if is_xlsx else read_xls(xls_path, log_path=log_path)
            # 按日期创建子文件夹
            date_dir = os.path.join(OUTPUT_DIR, date_str)
            os.makedirs(date_dir, exist_ok=True)
            output_path = os.path.join(date_dir, f"{date_str}.png")
            draw_table(data, date_str, output_path)
            # 品名明细 PNG
            try:
                items, _ = read_xlsx_items(xls_path, log_path=log_path) if is_xlsx else read_xls_items(xls_path, log_path=log_path)
                detail_path = os.path.join(date_dir, f"{date_str}_品名.png")
                draw_table_detail(items, date_str, detail_path)
            except Exception as detail_err:
                detail_log = log_path or os.path.join(OUTPUT_DIR, '汇总错误.log')
                show_error(f"品名明细生成失败: {detail_err}", detail_log, parent=root)
            os.startfile(output_path)
            status_var.set(f"已生成: {os.path.basename(output_path)} + 品名明细")
        except Exception as e:
            show_error(f"{type(e).__name__}: {e}", log_path, parent=root)
            status_var.set("")

    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', on_drop)

    root.mainloop()


# ==================== 核心逻辑 ====================
def _extract_date_from_xls(sh) -> str:
    """
    从 xls sheet 的 Row 2 提取日期范围, 返回用于显示的日期字符串.
    Row 2 格式: "日期:2026-06-03 00:00:00 至 2026-06-05 23:59:59"
    日期有滞后性: 结束日期 -1 天 = 实际最后完整统计日.
    - 单日 → "2026-06-03"
    - 跨天 → "2026-06-03至2026-06-04"
    提取失败返回 None.
    """
    try:
        row2_text = ' '.join([str(sh.cell_value(2, c)) for c in range(sh.ncols)])
        m = re.search(r'日期[:：]\s*(\d{4}-\d{2}-\d{2}).*?至.*?(\d{4}-\d{2}-\d{2})', row2_text)
        if m:
            start_str, end_str = m.group(1), m.group(2)
            # 当天导出当天数据: start == end, 无滞后, 直接返回
            if start_str == end_str:
                return start_str
            # 结束日期 -1 天 = 实际最后完整统计日
            end_date = datetime.strptime(end_str, '%Y-%m-%d')
            actual_end = end_date - timedelta(days=1)
            actual_end_str = actual_end.strftime('%Y-%m-%d')
            if start_str == actual_end_str:
                return start_str
            return f'{start_str}至{actual_end_str}'
    except Exception:
        pass
    return None


def _extract_date_from_xlsx(sh) -> str:
    """
    从 xlsx sheet 的 Row 3 (1-indexed) 提取日期范围.
    openpyxl 行列均为 1-indexed, Row 3 对应 xlrd 的 Row 2 (0-indexed).
    """
    try:
        row2_text = ' '.join([str(sh.cell(3, c).value or '') for c in range(1, sh.max_column + 1)])
        m = re.search(r'日期[:：]\s*(\d{4}-\d{2}-\d{2}).*?至.*?(\d{4}-\d{2}-\d{2})', row2_text)
        if m:
            start_str, end_str = m.group(1), m.group(2)
            if start_str == end_str:
                return start_str
            end_date = datetime.strptime(end_str, '%Y-%m-%d')
            actual_end = end_date - timedelta(days=1)
            actual_end_str = actual_end.strftime('%Y-%m-%d')
            if start_str == actual_end_str:
                return start_str
            return f'{start_str}至{actual_end_str}'
    except Exception:
        pass
    return None


def read_xls(path: str, log_path: str = None) -> tuple[dict, str]:
    """
    读 xls, 按品牌汇总销售金额.
    返回: ({品牌: 金额}, 日期字符串)  — 日期优先从 xls 内部提取, 失败则用文件名.
    """
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)

    # 从 xls 内部提取日期
    date_str = _extract_date_from_xls(sh)
    if not date_str:
        date_str = extract_date_from_filename(os.path.basename(path))

    # 定位表头
    header_row = brand_col = amount_col = None
    for i in range(min(20, sh.nrows)):  # 表头通常在前 20 行
        row = [str(sh.cell_value(i, c)).strip() for c in range(sh.ncols)]
        if any('品牌名称' in c for c in row) and any('销售金额' in c for c in row):
            header_row = i
            brand_col = next(c for c, v in enumerate(row) if '品牌名称' in v)
            amount_col = next(c for c, v in enumerate(row) if '销售金额' in v)
            break

    if header_row is None:
        raise ValueError("找不到 '品牌名称' 和 '销售金额' 列")

    # 汇总
    summary = defaultdict(float)
    skipped_count = 0
    for i in range(header_row + 1, sh.nrows):
        brand = str(sh.cell_value(i, brand_col)).strip()
        if not brand or brand.startswith('合计') or brand in ('总', '总计'):
            continue
        try:
            amount = float(sh.cell_value(i, amount_col))
        except (ValueError, TypeError):
            # 尝试清理千分位逗号后再转
            raw = str(sh.cell_value(i, amount_col)).strip().replace(',', '')
            try:
                amount = float(raw)
            except (ValueError, TypeError):
                skipped_count += 1
                continue
        if amount == 0:
            continue
        summary[brand] += amount

    if skipped_count > 0 and log_path:
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f'[{datetime.now()}] 警告: {skipped_count} 行因金额格式异常被跳过\n')
        except Exception:
            pass

    if not summary:
        raise ValueError("未提取到任何品牌销售数据")

    # 降序 + 保留 2 位小数
    data = {k: round(v, 2) for k, v in sorted(summary.items(), key=lambda x: -x[1])}
    return data, date_str


def read_xlsx(path: str, log_path: str = None) -> tuple[dict, str]:
    """
    读 xlsx, 按品牌汇总销售金额.
    返回: ({品牌: 金额}, 日期字符串)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active

    date_str = _extract_date_from_xlsx(sh)
    if not date_str:
        date_str = extract_date_from_filename(os.path.basename(path))

    # 定位表头 (openpyxl 1-indexed)
    header_row = brand_col = amount_col = None
    for i in range(1, min(21, sh.max_row + 1)):
        row = [str(sh.cell(i, c).value or '').strip() for c in range(1, sh.max_column + 1)]
        if any('品牌名称' in c for c in row) and any('销售金额' in c for c in row):
            header_row = i
            brand_col = next(c for c, v in enumerate(row) if '品牌名称' in v) + 1
            amount_col = next(c for c, v in enumerate(row) if '销售金额' in v) + 1
            break

    if header_row is None:
        raise ValueError("找不到 '品牌名称' 和 '销售金额' 列")

    summary = defaultdict(float)
    skipped_count = 0
    for i in range(header_row + 1, sh.max_row + 1):
        brand = str(sh.cell(i, brand_col).value or '').strip()
        if not brand or brand.startswith('合计') or brand in ('总', '总计'):
            continue
        try:
            amount = float(sh.cell(i, amount_col).value)
        except (ValueError, TypeError):
            raw = str(sh.cell(i, amount_col).value or '').strip().replace(',', '')
            try:
                amount = float(raw)
            except (ValueError, TypeError):
                skipped_count += 1
                continue
        if amount == 0:
            continue
        summary[brand] += amount

    if skipped_count > 0 and log_path:
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f'[{datetime.now()}] 警告: {skipped_count} 行因金额格式异常被跳过\n')
        except Exception:
            pass

    if not summary:
        raise ValueError("未提取到任何品牌销售数据")

    data = {k: round(v, 2) for k, v in sorted(summary.items(), key=lambda x: -x[1])}
    wb.close()
    return data, date_str


def read_xls_items(path: str, log_path: str = None) -> tuple[list, str]:
    """
    读 xls, 按品名逐行提取明细 (不汇总).
    返回: ([(品名, 品牌, 条码, 数量, 金额, 库存), ...], 日期字符串)
    保持原 xls 顺序, 过滤非数据行.
    """
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)

    date_str = _extract_date_from_xls(sh)
    if not date_str:
        date_str = extract_date_from_filename(os.path.basename(path))

    # 定位表头
    header_row = None
    item_col = brand_col = barcode_col = qty_col = amount_col = stock_col = None
    for i in range(min(20, sh.nrows)):
        row = [str(sh.cell_value(i, c)).strip() for c in range(sh.ncols)]
        if any('品牌名称' in c for c in row) and any('销售金额' in c for c in row):
            header_row = i
            brand_col = next(c for c, v in enumerate(row) if '品牌名称' in v)
            amount_col = next(c for c, v in enumerate(row) if '销售金额' in v)
            # 品名列
            for c, v in enumerate(row):
                if '品名' in v and '品牌' not in v:
                    item_col = c
                    break
            # 条码列
            for c, v in enumerate(row):
                if '条码' in v or '国际条码' in v:
                    barcode_col = c
                    break
            # 数量列
            for c, v in enumerate(row):
                if '销售数量' in v or '数量' in v:
                    qty_col = c
                    break
            # 库存列
            for c, v in enumerate(row):
                if '库存' in v:
                    stock_col = c
                    break
            break

    if header_row is None:
        raise ValueError("找不到 '品牌名称' 和 '销售金额' 列")
    if item_col is None:
        raise ValueError("找不到 '品名' 列")
    if qty_col is None:
        raise ValueError("找不到 '销售数量' 列")

    items = []
    skipped_count = 0
    for i in range(header_row + 1, sh.nrows):
        brand = str(sh.cell_value(i, brand_col)).strip()
        if not brand or brand.startswith('合计') or brand in ('总', '总计'):
            continue
        item_name = str(sh.cell_value(i, item_col)).strip()
        if not item_name:
            continue
        barcode = str(sh.cell_value(i, barcode_col)).strip() if barcode_col is not None else ''
        try:
            qty = int(float(sh.cell_value(i, qty_col)))
        except (ValueError, TypeError):
            qty = 0
        try:
            amount = float(sh.cell_value(i, amount_col))
        except (ValueError, TypeError):
            raw = str(sh.cell_value(i, amount_col)).strip().replace(',', '')
            try:
                amount = float(raw)
            except (ValueError, TypeError):
                skipped_count += 1
                continue
        if amount == 0:
            continue
        stock_raw = str(sh.cell_value(i, stock_col)).strip() if stock_col is not None else ''
        try:
            stock = int(float(stock_raw))
        except (ValueError, TypeError):
            stock = '-'
        items.append((item_name, brand, barcode, qty, round(amount, 2), stock))

    if skipped_count > 0 and log_path:
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f'[{datetime.now()}] 警告: {skipped_count} 行因金额格式异常被跳过 (品名明细)\n')
        except Exception:
            pass

    return items, date_str


def read_xlsx_items(path: str, log_path: str = None) -> tuple[list, str]:
    """
    读 xlsx, 按品名逐行提取明细 (不汇总).
    返回: ([(品名, 品牌, 条码, 数量, 金额, 库存), ...], 日期字符串)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active

    date_str = _extract_date_from_xlsx(sh)
    if not date_str:
        date_str = extract_date_from_filename(os.path.basename(path))

    # 定位表头 (openpyxl 1-indexed)
    header_row = None
    item_col = brand_col = barcode_col = qty_col = amount_col = stock_col = None
    for i in range(1, min(21, sh.max_row + 1)):
        row = [str(sh.cell(i, c).value or '').strip() for c in range(1, sh.max_column + 1)]
        if any('品牌名称' in c for c in row) and any('销售金额' in c for c in row):
            header_row = i
            brand_col = next(c for c, v in enumerate(row) if '品牌名称' in v) + 1
            amount_col = next(c for c, v in enumerate(row) if '销售金额' in v) + 1
            for c, v in enumerate(row):
                if '品名' in v and '品牌' not in v:
                    item_col = c + 1
                    break
            for c, v in enumerate(row):
                if '条码' in v or '国际条码' in v:
                    barcode_col = c + 1
                    break
            for c, v in enumerate(row):
                if '销售数量' in v or '数量' in v:
                    qty_col = c + 1
                    break
            for c, v in enumerate(row):
                if '库存' in v:
                    stock_col = c + 1
                    break
            break

    if header_row is None:
        raise ValueError("找不到 '品牌名称' 和 '销售金额' 列")
    if item_col is None:
        raise ValueError("找不到 '品名' 列")
    if qty_col is None:
        raise ValueError("找不到 '销售数量' 列")

    items = []
    skipped_count = 0
    for i in range(header_row + 1, sh.max_row + 1):
        brand = str(sh.cell(i, brand_col).value or '').strip()
        if not brand or brand.startswith('合计') or brand in ('总', '总计'):
            continue
        item_name = str(sh.cell(i, item_col).value or '').strip()
        if not item_name:
            continue
        barcode = str(sh.cell(i, barcode_col).value or '').strip() if barcode_col is not None else ''
        try:
            qty = int(float(sh.cell(i, qty_col).value))
        except (ValueError, TypeError):
            qty = 0
        try:
            amount = float(sh.cell(i, amount_col).value)
        except (ValueError, TypeError):
            raw = str(sh.cell(i, amount_col).value or '').strip().replace(',', '')
            try:
                amount = float(raw)
            except (ValueError, TypeError):
                skipped_count += 1
                continue
        if amount == 0:
            continue
        stock_raw = str(sh.cell(i, stock_col).value or '').strip() if stock_col is not None else ''
        try:
            stock = int(float(stock_raw))
        except (ValueError, TypeError):
            stock = '-'
        items.append((item_name, brand, barcode, qty, round(amount, 2), stock))

    if skipped_count > 0 and log_path:
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f'[{datetime.now()}] 警告: {skipped_count} 行因金额格式异常被跳过 (品名明细)\n')
        except Exception:
            pass

    wb.close()
    return items, date_str


def draw_table(data: dict, date_str: str, output_path: str):
    """Pillow 画表格, 保存为 PNG"""
    if not os.path.exists(FONT_PATH):
        raise FileNotFoundError(f"找不到中文字体: {FONT_PATH}")

    font_title = ImageFont.truetype(FONT_PATH, SIZE_TITLE)
    font_header = ImageFont.truetype(FONT_PATH, SIZE_HEADER)
    font_body = ImageFont.truetype(FONT_PATH, SIZE_BODY)
    font_total = ImageFont.truetype(FONT_PATH, SIZE_TOTAL)

    total_amount = round(sum(data.values()), 2)
    n = len(data)

    # ---- 自适应宽度 ----
    # 计算标题宽度
    title = f"销售汇总  {date_str}"
    title_w = font_title.getbbox(title)[2] - font_title.getbbox(title)[0]

    # 计算最宽品牌名宽度
    max_brand_w = 0
    for brand in data:
        bw = font_body.getbbox(brand)[2] - font_body.getbbox(brand)[0]
        if bw > max_brand_w:
            max_brand_w = bw

    brand_col_w = max(COL_BRAND_W, max_brand_w + 40)  # 最少 260, 品牌名宽则扩展
    amount_col_w = COL_AMOUNT_W
    width = max(brand_col_w + amount_col_w, title_w) + 2 * PADDING

    height = (n + 2) * ROW_H + 2 * PADDING + 50  # +2 表头+总计, +50 标题

    img = Image.new('RGB', (width, height), COLOR_BG)
    draw = ImageDraw.Draw(img)

    # ---- 标题 (居中) ----
    draw.text(((width - title_w) / 2, PADDING / 2), title,
              font=font_title, fill=COLOR_TEXT)

    # ---- 表头 ----
    x_left = PADDING
    x_mid = PADDING + brand_col_w
    x_right = PADDING + brand_col_w + amount_col_w
    y = PADDING + 50

    draw.rectangle([x_left, y, x_right, y + ROW_H], outline=COLOR_BORDER, width=2)
    # 垂直分隔线
    draw.line([(x_mid, y), (x_mid, y + ROW_H)], fill=COLOR_BORDER, width=2)
    draw.text((x_left + 20, y + 12), "品牌名称", font=font_header, fill=COLOR_TEXT)
    draw.text((x_mid + 20, y + 12), "销售金额（元）", font=font_header, fill=COLOR_TEXT)

    # ---- 数据行 ----
    y += ROW_H
    for brand, amount in data.items():
        draw.rectangle([x_left, y, x_right, y + ROW_H],
                       outline=COLOR_BORDER, width=1)
        draw.line([(x_mid, y), (x_mid, y + ROW_H)], fill=COLOR_BORDER, width=1)
        brand_text = brand
        brand_max_w = brand_col_w - 40
        if draw.textlength(brand, font=font_body) > brand_max_w:
            while len(brand_text) > 0 and draw.textlength(brand_text + '…', font=font_body) > brand_max_w:
                brand_text = brand_text[:-1]
            brand_text += '…'
        draw.text((x_left + 20, y + 14), brand_text, font=font_body, fill=COLOR_TEXT)
        draw.text((x_mid + 20, y + 14), f"{amount:.2f}", font=font_body, fill=COLOR_TEXT)
        y += ROW_H

    # ---- 总计行（加粗 + 双线边框） ----
    draw.rectangle([x_left, y, x_right, y + ROW_H],
                   outline=COLOR_BORDER, width=3)
    draw.line([(x_mid, y), (x_mid, y + ROW_H)], fill=COLOR_BORDER, width=3)
    draw.text((x_left + 20, y + 12), "总销售金额", font=font_total, fill=COLOR_TEXT)
    draw.text((x_mid + 20, y + 12), f"{total_amount:.2f}", font=font_total, fill=COLOR_TEXT)

    img.save(output_path, 'PNG')


def draw_table_detail(items: list, date_str: str, output_path: str):
    """Pillow 画品名明细表格 (7 列: 行号/条码/品名/品牌/数量/金额/库存), 保存为 PNG"""
    if not os.path.exists(FONT_PATH):
        raise FileNotFoundError(f"找不到中文字体: {FONT_PATH}")

    font_title = ImageFont.truetype(FONT_PATH, SIZE_TITLE)
    font_header = ImageFont.truetype(FONT_PATH, SIZE_HEADER)
    font_body = ImageFont.truetype(FONT_PATH, SIZE_BODY)
    font_total = ImageFont.truetype(FONT_PATH, SIZE_TOTAL)

    n = len(items)
    total_qty = sum(qty for _, _, _, qty, _, _ in items)
    total_amount = round(sum(amt for _, _, _, _, amt, _ in items), 2)

    # ---- 自适应宽度 ----
    title = f"销售明细  {date_str}"
    title_w = font_title.getbbox(title)[2] - font_title.getbbox(title)[0]

    max_item_w = 0
    max_brand_w = 0
    for item_name, brand, _, _, _, _ in items:
        iw = font_body.getbbox(item_name)[2] - font_body.getbbox(item_name)[0]
        bw = font_body.getbbox(brand)[2] - font_body.getbbox(brand)[0]
        if iw > max_item_w:
            max_item_w = iw
        if bw > max_brand_w:
            max_brand_w = bw

    num_col_w = DETAIL_COL_NUM_W
    barcode_col_w = DETAIL_COL_BARCODE_W
    item_col_w = max(DETAIL_COL_ITEM_W, max_item_w + 40)
    brand_col_w = max(DETAIL_COL_BRAND_W, max_brand_w + 40)
    qty_col_w = DETAIL_COL_QTY_W
    amount_col_w = DETAIL_COL_AMOUNT_W
    stock_col_w = DETAIL_COL_STOCK_W
    total_col_w = num_col_w + barcode_col_w + item_col_w + brand_col_w + qty_col_w + amount_col_w + stock_col_w
    width = max(total_col_w, title_w) + 2 * PADDING

    height = (n + 2) * DETAIL_ROW_H + 2 * PADDING + 50

    img = Image.new('RGB', (width, height), COLOR_BG)
    draw = ImageDraw.Draw(img)

    # ---- 标题 (居中) ----
    draw.text(((width - title_w) / 2, PADDING / 2), title,
              font=font_title, fill=COLOR_TEXT)

    # ---- 表头 ----
    x1 = PADDING
    x2 = x1 + num_col_w
    x3 = x2 + barcode_col_w
    x4 = x3 + item_col_w
    x5 = x4 + brand_col_w
    x6 = x5 + qty_col_w
    x7 = x6 + amount_col_w
    x_right = x7 + stock_col_w
    y = PADDING + 50

    draw.rectangle([x1, y, x_right, y + DETAIL_ROW_H], outline=COLOR_BORDER, width=2)
    for x in [x2, x3, x4, x5, x6, x7]:
        draw.line([(x, y), (x, y + DETAIL_ROW_H)], fill=COLOR_BORDER, width=2)
    draw.text((x1 + 10, y + 12), "行号", font=font_header, fill=COLOR_TEXT)
    draw.text((x2 + 10, y + 12), "国际条码", font=font_header, fill=COLOR_TEXT)
    draw.text((x3 + 20, y + 12), "品名", font=font_header, fill=COLOR_TEXT)
    draw.text((x4 + 20, y + 12), "品牌", font=font_header, fill=COLOR_TEXT)
    draw.text((x5 + 10, y + 12), "数量", font=font_header, fill=COLOR_TEXT)
    draw.text((x6 + 10, y + 12), "金额（元）", font=font_header, fill=COLOR_TEXT)
    draw.text((x7 + 10, y + 12), "库存", font=font_header, fill=COLOR_TEXT)

    # ---- 数据行 ----
    y += DETAIL_ROW_H
    for idx, (item_name, brand, barcode, qty, amount, stock) in enumerate(items, 1):
        draw.rectangle([x1, y, x_right, y + DETAIL_ROW_H],
                       outline=COLOR_BORDER, width=1)
        for x in [x2, x3, x4, x5, x6, x7]:
            draw.line([(x, y), (x, y + DETAIL_ROW_H)], fill=COLOR_BORDER, width=1)

        # 行号
        draw.text((x1 + 10, y + 14), str(idx), font=font_body, fill=COLOR_TEXT)

        # 条码列截断
        barcode_text = barcode
        barcode_max_w = barcode_col_w - 20
        if draw.textlength(barcode, font=font_body) > barcode_max_w:
            while len(barcode_text) > 0 and draw.textlength(barcode_text + '…', font=font_body) > barcode_max_w:
                barcode_text = barcode_text[:-1]
            barcode_text += '…'
        draw.text((x2 + 10, y + 14), barcode_text, font=font_body, fill=COLOR_TEXT)

        # 品名列截断
        item_text = item_name
        item_max_w = item_col_w - 40
        if draw.textlength(item_name, font=font_body) > item_max_w:
            while len(item_text) > 0 and draw.textlength(item_text + '…', font=font_body) > item_max_w:
                item_text = item_text[:-1]
            item_text += '…'
        draw.text((x3 + 20, y + 14), item_text, font=font_body, fill=COLOR_TEXT)

        # 品牌列截断
        brand_text = brand
        brand_max_w = brand_col_w - 40
        if draw.textlength(brand, font=font_body) > brand_max_w:
            while len(brand_text) > 0 and draw.textlength(brand_text + '…', font=font_body) > brand_max_w:
                brand_text = brand_text[:-1]
            brand_text += '…'
        draw.text((x4 + 20, y + 14), brand_text, font=font_body, fill=COLOR_TEXT)

        draw.text((x5 + 10, y + 14), str(qty), font=font_body, fill=COLOR_TEXT)
        draw.text((x6 + 10, y + 14), f"{amount:.2f}", font=font_body, fill=COLOR_TEXT)
        draw.text((x7 + 10, y + 14), str(stock), font=font_body, fill=COLOR_TEXT)
        y += DETAIL_ROW_H

    # ---- 总计行 ----
    draw.rectangle([x1, y, x_right, y + DETAIL_ROW_H],
                   outline=COLOR_BORDER, width=3)
    for x in [x2, x3, x4, x5, x6, x7]:
        draw.line([(x, y), (x, y + DETAIL_ROW_H)], fill=COLOR_BORDER, width=3)
    draw.text((x1 + 10, y + 12), "合计", font=font_total, fill=COLOR_TEXT)
    # 条码/品名/品牌列空
    draw.text((x5 + 10, y + 12), str(total_qty), font=font_total, fill=COLOR_TEXT)
    draw.text((x6 + 10, y + 12), f"{total_amount:.2f}", font=font_total, fill=COLOR_TEXT)

    img.save(output_path, 'PNG')


# ==================== 入口 ====================
def main():
    if len(sys.argv) < 2:
        launch_gui()
        return

    xls_path = os.path.abspath(sys.argv[1])
    log_path = os.path.join(os.path.dirname(xls_path) or '.', '汇总错误.log')

    if not os.path.exists(xls_path):
        show_error(f"文件不存在:\n{xls_path}", log_path)
        return

    if not (xls_path.lower().endswith('.xls') or xls_path.lower().endswith('.xlsx')):
        show_error(f"不支持的文件格式，请拖入 .xls 或 .xlsx 文件:\n{xls_path}", log_path)
        return

    if not os.path.isdir(OUTPUT_DIR):
        show_error(f"输出目录不存在:\n{OUTPUT_DIR}\n请先创建该目录。", log_path)
        return

    is_xlsx = xls_path.lower().endswith('.xlsx')
    try:
        data, date_str = read_xlsx(xls_path, log_path=log_path) if is_xlsx else read_xls(xls_path, log_path=log_path)
        # 按日期创建子文件夹
        date_dir = os.path.join(OUTPUT_DIR, date_str)
        os.makedirs(date_dir, exist_ok=True)
        output_path = os.path.join(date_dir, f"{date_str}.png")
        draw_table(data, date_str, output_path)
        # 品名明细 PNG
        try:
            items, _ = read_xlsx_items(xls_path, log_path=log_path) if is_xlsx else read_xls_items(xls_path, log_path=log_path)
            detail_path = os.path.join(date_dir, f"{date_str}_品名.png")
            draw_table_detail(items, date_str, detail_path)
        except Exception as detail_err:
            detail_log = log_path or os.path.join(OUTPUT_DIR, '汇总错误.log')
            show_error(f"品名明细生成失败: {detail_err}", detail_log)
        # 自动用默认看图工具打开品牌汇总
        os.startfile(output_path)
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        show_error(err, log_path)


if __name__ == '__main__':
    main()
